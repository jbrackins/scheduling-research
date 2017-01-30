from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import re
import sys
import datetime as dt
import dateutil.parser as dparser
import collections
import classrooms
import os
import platform
import textwrap as tw

# if a class doesn't meet, empty cell is returned.

no_days = "       "
# dict containing all possible class times
class_days  = {}
class_days2  = {}
# dict containing all possible class times
class_times  = {}

courses      = {}
rooms_used   = {}

# list of buildings
buildings    = {}

# list of rooms for a given building
rooms        = {"Monday": {}, "Tuesday" : {}, "Wednesday" : {}, "Thursday" : {}, "Friday" : {}, "Saturday" : {}}
room_caps    = {"Monday": {}, "Tuesday" : {}, "Wednesday" : {}, "Thursday" : {}, "Friday" : {}, "Saturday" : {}}


school_day = {"Monday": {}, "Tuesday" : {}, "Wednesday" : {}, "Thursday" : {}, "Friday" : {}, "Saturday" : {}}
def build_path(yr, sem, location):
    old_dir = os.getcwd()

    if not os.path.exists('./dump'):
        os.mkdir("dump")
    os.chdir("dump")
    if not os.path.exists(yr):
        os.mkdir(yr)
    os.chdir(yr)
    if not os.path.exists(sem):
        os.mkdir(sem)
    os.chdir(sem)
    if not os.path.exists(location):
        os.mkdir(location)
    os.chdir(old_dir)

def fill_timeline(x):
    for time in classrooms.course_times:
        time = str(time)
        if time not in x:
            x[time] = 0
    return x

def get_start_time(time):
    return time.split("-")[0]

def get_building(build_room):
    return build_room.split(" ")[0]

def get_military_time(time):

    # get rid of that weird extra colon in the data...
    if time:
        time = time[::-1].replace(":", ""[::-1], 1)[::-1]
        date=dparser.parse(time)
        return date.strftime('%H:%M')
    else:
        return None

def get_semester(filename):
    # Determine the semester statistics (Spring/Fall) and Year

    # todo: add search for summer as well
    semester = None
    year     = None

    # find the year
    year = re.findall(r'\d+', filename)
    year = year[0]

    # find the semester
    if "FA" in filename:
        semester = "Fall"
    if "SP" in filename:
        semester = "Spring"

    return semester, year



if __name__ == "__main__":
    if(len(sys.argv) > 1):
        filename =  sys.argv[1] 
        wb = load_workbook(filename)
        ws = wb.active
        for i in range(2, len(ws.rows)):
            index_subject       = "C"  + str(i)
            index_course_number = "D"  + str(i)

            index_status        = "J"  + str(i)
            index_days_of_week  = "AP" + str(i)
            index_start_time    = "AH" + str(i)
            index_end_time      = "AI" + str(i)
            index_capacity      = "AG" + str(i)

            index_building = "AC" + str(i)
            index_room     = "AD" + str(i)
            index_course = "I" + str(i)
            index_count  = "M" + str(i)

            status = str(ws[index_status].value)

            # We only care about active courses...
            if "A" in status:

                subject = ws[index_subject].value
                course_number = ws[index_course_number].value
                course = ws[index_course].value

                course_full = str(subject) + " " + str(course_number) + ": " + str(course)
                count  = ws[index_count].value

                capacity = ws[index_capacity].value
                if capacity is None or capacity < 1:
                    capacity = float(1)
                else:
                    capacity = float(capacity)
                building = ws[index_building].value
                room     = ws[index_room].value
                location = str(building) + " " + str(room)

                days_of_week = str(ws[index_days_of_week].value).replace("Th", "R")
                start_time   = ws[index_start_time].value
                if start_time is None:
                    start_time = "None"
                else:
                    start_time   = get_military_time( str(start_time) )

                end_time     = ws[index_end_time].value

                if end_time is None:
                    end_time = "None"
                else:
                    end_time   = get_military_time( str(end_time) )


                course_times  = str(start_time) + "-" + str(end_time)

                if days_of_week != no_days: 
                    if days_of_week not in class_days:
                        class_days[days_of_week] = {}
                        class_days2[days_of_week] = {}

                    if course_times not in class_days[days_of_week]:
                        class_days[days_of_week][course_times] = {}
                        class_days2[days_of_week][course_times] = {}

                    if course_full not in class_days[days_of_week][course_times]:
                        class_days[days_of_week][course_times][course_full] = {location : (count/capacity) * 100 }
                        class_days2[days_of_week][course_times][course_full] = {location : capacity }


            #if course_date not in class_times:
            #    class_times[course_date] = {}

    for x in ["M", "T", "W", "R", "F", "S"]:
        d = {}
        t = {}
        c = {}
        #print(classrooms.day[x] + ":")
        for i in sorted(class_days.keys()):
            if x in i:
                # Days of the week in which course occurs
                #print("\t" + i.replace(" ", "") + ":")
                for j in sorted(class_days[i].keys()):
                    # Course Times
                    #print("\t\t" + j)
                    for k in sorted(class_days[i][j].keys()):
                        # Course Name
                        #print("\t\t\t" + k)
                        for l in class_days[i][j][k].keys():
                            # Course Location
                            #print("\t\t\t  " + l)
                            if class_days[i][j][k][l] >= 100:
                                cnt = 100
                            else:
                                cnt     = class_days[i][j][k][l]
                            
                            max_cap = class_days2[i][j][k][l]
                            #print("\t\t\t    " + cnt )

                            # create entry into dictionary
                            y = classrooms.day[x]
                            j2 = get_start_time(j)
                            if l not in rooms[y]:
                                rooms[y][l] = {}
                                room_caps[y][l] = {}
                            if k not in rooms[y][l]:
                                rooms[y][l][k] = {}
                            if j2 not in rooms[y][l][k]:
                                rooms[y][l][k][j2] = {}
                            rooms[y][l][k][j2] = cnt
                            room_caps[y][l] = max_cap
                            #school_day[y][j2][k][l] = cnt



    #print(class_times)
    for weekday in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]:
        #print(weekday,":")
        for location in sorted(rooms[weekday].keys()):
            #print("\t",location)
            #for course in sorted(rooms[weekday][location].keys()):
                #print("\t\t", course)
                #if course in classrooms.main_rooms:
                #for time in sorted(rooms[weekday][location][course].keys()):
                    #print("\t\t\t", time)
                    #print("\t\t\t\t", rooms[weekday][location][course][time])
            
            if "None" not in location and get_building(location) in classrooms.main_rooms:
                a = rooms[weekday][location]
                r = room_caps[weekday][location]
                l = []
                d = {}
                #print(a.keys())
                for b in a.values():
                    l.append(b)
                for c in l:

                    for i, j in c.items():
                        # if there is already an entry for this given time,
                        # the class is double-listed in the catalog. combine stats.
                        if i in d:
                            d[i] = d[i] + j
                        else:
                            d[i] = j

                #print(d)

                d = fill_timeline(d)
                plot = collections.OrderedDict(sorted(d.items()))

                
                if len(plot) > 0:
                    #print(plot)
                    sem, yr = get_semester(filename)
                    room_title = "Room Usage Statistics for "
                    room_title = room_title + location + " on " + weekday + "s"
                    room_title = room_title + " - " + sem + " " + yr
                    plt.figure(figsize=(20,10))

                    plt.title(room_title)
                    plt.xlabel('Time')
                    y_lab = 'Room Utilization (in percent)'
                    plt.ylabel(y_lab)
                    # Text places anywhere within the Axis
                    plt.text(0.6, 90, location + " capacity is " + str(int(r)),
                        horizontalalignment='right', backgroundcolor='pink')
                    plt.bar(range(len(plot)), plot.values(), align='center')
                    plt.xticks(range(len(plot)), list(plot.keys()))
                    plt.axis(ymin = 0, ymax= 100)
                    plt.xticks(rotation=70)

                    #plt.show()

                    build_path(yr, sem, location.replace(" ","_"))
                    plt_file = "dump/"  + yr + "/" + sem  + "/" + location.replace(" ","_")  + "/" + weekday + "_" + location.replace(" ", "_") 
                    #print(plt_file)
                    plt.savefig(plt_file, format='pdf', bbox_inches='tight')
                    plt.clf()        
                    plt.close() 








