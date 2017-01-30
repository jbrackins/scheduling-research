from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import re
import sys
import datetime as dt
import dateutil.parser as dparser
import collections
import classrooms

import platform

# if a class doesn't meet, empty cell is returned.

no_days = "       "
# dict containing all possible class times
class_days  = {}
# dict containing all possible class times
class_times  = {}

courses      = {}
rooms_used   = {}

# list of buildings
buildings    = {}

# list of rooms for a given building
rooms        = {}


school_day = {"Monday": {}, "Tuesday" : {}, "Wednesday" : {}, "Thursday" : {}, "Friday" : {}, "Saturday" : {}}

def get_start_time(time):
    return time.split("-")[0]

def get_building(build_room):
    return build_room.split(" ")[0]

def get_military_time(time):

    # get rid of that weird extra colon in the data...
    if time:
        time = time[::-1].replace(":", ""[::-1], 1)[::-1]
        date=dparser.parse(time)
        return date.strftime('%H:%M')
    else:
        return None

def get_semester(filename):
    # Determine the semester statistics (Spring/Fall) and Year

    # todo: add search for summer as well
    semester = None
    year     = None

    # find the year
    year = re.findall(r'\d+', filename)
    year = year[0]

    # find the semester
    if "FA" in filename:
        semester = "Fall"
    if "SP" in filename:
        semester = "Spring"

    return semester, year



if __name__ == "__main__":
    if(len(sys.argv) > 1):
        filename =  sys.argv[1] 
        wb = load_workbook(filename)
        ws = wb.active
        for i in range(2, len(ws.rows)):
            index_subject       = "C"  + str(i)
            index_course_number = "D"  + str(i)

            index_status        = "J"  + str(i)
            index_days_of_week  = "AP" + str(i)
            index_start_time    = "AH" + str(i)
            index_end_time      = "AI" + str(i)
            index_capacity      = "AG" + str(i)

            index_building = "AC" + str(i)
            index_room     = "AD" + str(i)
            index_course = "I" + str(i)
            index_count  = "M" + str(i)

            status = str(ws[index_status].value)

            # We only care about active courses...
            if "A" in status:

                subject = ws[index_subject].value
                course_number = ws[index_course_number].value
                course = ws[index_course].value

                course_full = str(subject) + " " + str(course_number) + ": " + str(course)
                count  = ws[index_count].value

                capacity = ws[index_capacity].value
                if capacity is None or capacity < 1:
                    capacity = float(1)
                else:
                    capacity = float(capacity)
                building = ws[index_building].value
                room     = ws[index_room].value
                location = str(building) + " " + str(room)

                days_of_week = str(ws[index_days_of_week].value).replace("Th", "R")
                start_time   = ws[index_start_time].value
                if start_time is None:
                    start_time = "None"
                else:
                    start_time   = get_military_time( str(start_time) )

                end_time     = ws[index_end_time].value

                if end_time is None:
                    end_time = "None"
                else:
                    end_time   = get_military_time( str(end_time) )


                course_times  = str(start_time) + "-" + str(end_time)

                if days_of_week != no_days: 
                    if days_of_week not in class_days:
                        class_days[days_of_week] = {}

                    if course_times not in class_days[days_of_week]:
                        class_days[days_of_week][course_times] = {}

                    if course_full not in class_days[days_of_week][course_times]:
                        class_days[days_of_week][course_times][course_full] = {location : (count/capacity) * 100 }


            #if course_date not in class_times:
            #    class_times[course_date] = {}

    for x in ["M", "T", "W", "R", "F", "S"]:
        d = {}
        t = {}
        c = {}
        #print(classrooms.day[x] + ":")
        for i in sorted(class_days.keys()):
            if x in i:
                # Days of the week in which course occurs
                #print("\t" + i.replace(" ", "") + ":")
                for j in sorted(class_days[i].keys()):
                    # Course Times
                    #print("\t\t" + j)
                    for k in sorted(class_days[i][j].keys()):
                        # Course Name
                        #print("\t\t\t" + k)
                        for l in class_days[i][j][k].keys():
                            # Course Statistics
                            #print("\t\t\t  " + l)
                            if class_days[i][j][k][l] >= 100:
                                cnt = 100
                            else:
                                cnt = class_days[i][j][k][l]
                            #print("\t\t\t    " + cnt )

                            # create entry into dictionary
                            y = classrooms.day[x]
                            j2 = get_start_time(j)
                            print(j2)
                            if j2 not in school_day[y]:
                                school_day[y][j2] = {}
                            if k not in school_day[y][j2]:
                                school_day[y][j2][k] = {}
                            if l not in school_day[y][j2][k]:
                                school_day[y][j2][k][l] = {}
                            school_day[y][j2][k][l] = cnt



    #print(class_times)
    for weekday in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]:
        print(weekday,":")
        for time in sorted(school_day[weekday].keys()):
            print("\t",time)
            for course in sorted(school_day[weekday][time].keys()):
                print("\t\t", course)
                #if course in classrooms.main_rooms:
                for build_room in sorted(school_day[weekday][time][course].keys()):
                    print("\t\t\t", build_room)
                    print("\t\t\t\t", school_day[weekday][time][course][build_room])
            
            a = school_day[weekday][time]
            l = []
            d = {}
            for b in a.values():
                l.append(b)
            for c in l:
                for i, j in c.items():
                    # Filter out courses that don't have rooms 
                    # and courses in buildings that don't matter
                    if "None" not in i and get_building(i) in classrooms.main_rooms:
                        d[i] = j

            print(d)


            plot = collections.OrderedDict(sorted(d.items()))
            
            if len(plot) > 0:
                print(plot)
                sem, yr = get_semester(filename)
                room_title = "Room Usage Statistics for "
                room_title = room_title + weekday + " (" + time + ")"

                room_title = room_title + " - " + sem + " " + yr
                plt.figure(figsize=(20,10))

                plt.title(room_title)
                plt.xlabel('Room')
                plt.ylabel('Room Utilization (in percent)')
                plt.bar(range(len(plot)), plot.values(), align='center')
                plt.xticks(range(len(plot)), list(plot.keys()))
                plt.xticks(rotation=70)

                #plt.show()

                plt_file = "dump/" + weekday + "_" + time.replace(":","_") + "_" + yr + "_" + sem
                print(plt_file)
                plt.savefig(plt_file, format='pdf', bbox_inches='tight')
                plt.clf()        







