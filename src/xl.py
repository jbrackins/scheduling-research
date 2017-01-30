from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import re
import sys

import classrooms

import platform

courses      = {}
rooms_used   = {}

# list of buildings
buildings    = {}

# list of rooms for a given building
rooms        = {}


def get_semester(filename):
    # Determine the semester statistics (Spring/Fall) and Year

    # todo: add search for summer as well
    semester = None
    year     = None

    # find the year
    year = re.findall(r'\d+', filename)
    year = year[0]

    # find the semester
    if "FA" in filename:
        semester = "Fall"
    if "SP" in filename:
        semester = "Spring"

    return semester, year



if __name__ == "__main__":
    if(len(sys.argv) > 1):
        filename =  sys.argv[1] 
        wb = load_workbook(filename)
        ws = wb.active
        for i in range(2, len(ws.rows)):
            index_building = "AC" + str(i)
            index_room     = "AD" + str(i)
            index_course = "I" + str(i)
            index_count  = "M" + str(i)
            course = ws[index_course].value
            count  = ws[index_count].value

            building = ws[index_building].value
            room     = ws[index_room].value
            location = str(building) + " " + str(room)

            if building not in buildings:
                # add new building to dictionary
                buildings[building] = {}

            #print(room in buildings[building])
            if room in buildings[building]:
                buildings[building][room] = buildings[building][room] + 1
            else:
                buildings[building][room] = 1


            # tally of the number of courses that use this room
            #if location in buildings[rooms[location]]:
            #    buildings[rooms[location]] = buildings[rooms[location]] + 1
            #else:
            #    buildings[rooms[location]] = 1

            #if course in courses:
            #    courses[course] = courses[course] + count 
            #else:
            #    courses[course] = count

    del buildings[None]
    del buildings["COFC"]
    #for course, count in courses.items():
    #    print(course, count)


    for loc, value in buildings.items():
        #print(loc)
        plot = buildings[loc]
        sem, yr = get_semester(filename)
        room_title = "Room Usage for "
        if loc in classrooms.code:
            room_title = room_title + classrooms.code[loc] + " (" + loc + ")"
        else:
            room_title = room_title + " (" + loc + ")"
        room_title = room_title + " - " + sem + " " + yr
        plt.figure(figsize=(20,10))

        plt.title(room_title)
        plt.xlabel('Room')
        plt.ylabel('Number of Courses Using Room')
        plt.bar(range(len(plot)), plot.values(), align='center')
        plt.xticks(range(len(plot)), list(plot.keys()))
        #plt.show()

        plt_file = loc + "_" + yr + "_" + sem
        print(plt_file)
        plt.savefig(plt_file, format='pdf', bbox_inches='tight')
        plt.clf()
        #print(buildings[loc])
        #if loc in classrooms.code:
        #    print(classrooms.code[loc], "(" + loc + ")")
        #else:
        #    print("Unknown Location", "(" + loc + ")")
        #for i, ct in value.items():
        #    print("\t",i, ct)
        #for room in loc.items():
        #    print(room)


    #D = {u'Label1':26, u'Label2': 17, u'Label3':30}

    # Remove courses with no room assigned
    #del buildings[rooms['None None']]
    #plt.bar(range(len(rooms_used)), rooms_used.values(), align='center')
    #plt.xticks(range(len(rooms_used)), list(rooms_used.keys()))

    #plt.show()




